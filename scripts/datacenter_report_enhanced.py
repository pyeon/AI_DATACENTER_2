import yfinance as yf
import pandas as pd
import requests
import os
from datetime import datetime
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

print("="*70)
print("📊 데이터센터 투자 자동화 시스템 v2.0 (Enhanced)")
print("="*70 + "\n")

# 환경변수에서 텔레그램 설정 읽기
TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN')
TELEGRAM_CHAT_ID = os.environ.get('TELEGRAM_CHAT_ID')

# 종목 리스트 (세부영역 추가)
STOCKS = [
    # AI 칩셋
    {'name': 'NVIDIA', 'ticker': 'NVDA', 'category': 'AI 인프라', 'sector': 'AI칩', 'sub_sector': 'GPU'},
    {'name': 'AMD', 'ticker': 'AMD', 'category': 'AI 인프라', 'sector': 'AI칩', 'sub_sector': 'GPU'},
    {'name': 'Intel', 'ticker': 'INTC', 'category': 'AI 인프라', 'sector': 'AI칩', 'sub_sector': 'CPU'},
    
    # AI 서버
    {'name': 'Super Micro', 'ticker': 'SMCI', 'category': 'AI 인프라', 'sector': 'AI서버', 'sub_sector': '서버제조'},
    {'name': 'Dell', 'ticker': 'DELL', 'category': 'AI 인프라', 'sector': 'AI서버', 'sub_sector': '서버제조'},
    
    # 전력/쿨링
    {'name': 'Vertiv', 'ticker': 'VRT', 'category': '전력/쿨링', 'sector': '전력', 'sub_sector': '전력관리'},
    {'name': 'Eaton', 'ticker': 'ETN', 'category': '전력/쿨링', 'sector': '전력', 'sub_sector': '전력관리'},
    {'name': 'LS ELECTRIC', 'ticker': '010120.KS', 'category': '전력/쿨링', 'sector': '전력', 'sub_sector': '전력기기'},
    {'name': 'Cummins', 'ticker': 'CMI', 'category': '전력/쿨링', 'sector': '발전', 'sub_sector': '발전기'},
    {'name': 'Generac', 'ticker': 'GNRC', 'category': '전력/쿨링', 'sector': '발전', 'sub_sector': '발전기'},
    {'name': 'Johnson Controls', 'ticker': 'JCI', 'category': '전력/쿨링', 'sector': '쿨링', 'sub_sector': 'HVAC'},
    {'name': 'Trane Tech', 'ticker': 'TT', 'category': '전력/쿨링', 'sector': '쿨링', 'sub_sector': 'HVAC'},
    
    # 네트워크
    {'name': 'Arista Networks', 'ticker': 'ANET', 'category': '네트워크', 'sector': '네트워크', 'sub_sector': '스위치'},
    {'name': 'Broadcom', 'ticker': 'AVGO', 'category': '네트워크', 'sector': '네트워크', 'sub_sector': '네트워크칩'},
    {'name': 'Marvell', 'ticker': 'MRVL', 'category': '네트워크', 'sector': '네트워크', 'sub_sector': '네트워크칩'},
    {'name': 'HFR', 'ticker': '230240.KQ', 'category': '네트워크', 'sector': '광통신', 'sub_sector': '광트랜시버'},
    {'name': 'Corning', 'ticker': 'GLW', 'category': '네트워크', 'sector': '광섬유', 'sub_sector': '광섬유케이블'},
    {'name': 'Lumentum', 'ticker': 'LITE', 'category': '네트워크', 'sector': '광통신', 'sub_sector': '광학부품'},
    
    # 메모리/스토리지
    {'name': 'SK hynix', 'ticker': '000660.KS', 'category': '메모리/스토리지', 'sector': 'HBM', 'sub_sector': 'HBM메모리'},
    {'name': 'Samsung', 'ticker': '005930.KS', 'category': '메모리/스토리지', 'sector': 'HBM', 'sub_sector': 'HBM메모리'},
    {'name': 'Micron', 'ticker': 'MU', 'category': '메모리/스토리지', 'sector': 'HBM', 'sub_sector': 'HBM메모리'},
    {'name': '한미반도체', 'ticker': '042700.KQ', 'category': '메모리/스토리지', 'sector': '패키징', 'sub_sector': '반도체패키징'},
    {'name': 'Amkor', 'ticker': 'AMKR', 'category': '메모리/스토리지', 'sector': '패키징', 'sub_sector': '반도체패키징'},
    {'name': 'Western Digital', 'ticker': 'WDC', 'category': '메모리/스토리지', 'sector': 'SSD', 'sub_sector': '스토리지'},
    
    # 데이터센터 REIT
    {'name': 'Digital Realty', 'ticker': 'DLR', 'category': 'DC 부동산', 'sector': 'DC REIT', 'sub_sector': '데이터센터REIT'},
    {'name': 'Equinix', 'ticker': 'EQIX', 'category': 'DC 부동산', 'sector': 'DC REIT', 'sub_sector': '데이터센터REIT'},
]

print(f"📋 총 {len(STOCKS)}개 종목 모니터링\n")

def calculate_rsi(prices, period=14):
    """RSI 계산 (0~100)"""
    try:
        delta = prices.diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return rsi.iloc[-1] if not rsi.empty else 50
    except:
        return 50

def calculate_macd(prices):
    """MACD 계산"""
    try:
        ema12 = prices.ewm(span=12, adjust=False).mean()
        ema26 = prices.ewm(span=26, adjust=False).mean()
        macd = ema12 - ema26
        signal = macd.ewm(span=9, adjust=False).mean()
        histogram = macd - signal
        return macd.iloc[-1], signal.iloc[-1], histogram.iloc[-1]
    except:
        return 0, 0, 0

def calculate_bollinger_bands(prices, period=20):
    """볼린저 밴드 계산"""
    try:
        sma = prices.rolling(window=period).mean()
        std = prices.rolling(window=period).std()
        upper = sma + (std * 2)
        lower = sma - (std * 2)
        current = prices.iloc[-1]
        bb_position = ((current - lower.iloc[-1]) / (upper.iloc[-1] - lower.iloc[-1])) * 100
        return bb_position if not pd.isna(bb_position) else 50
    except:
        return 50

def get_trading_signal(data):
    """매매 신호 종합 판단"""
    signals = []
    
    # 단기 신호
    if data['rsi'] < 30:
        signals.append('단기과매도')
    elif data['rsi'] > 70:
        signals.append('단기과매수')
    
    # 중기 신호
    if data['macd_histogram'] > 0 and data['golden_cross']:
        signals.append('중기상승')
    elif data['macd_histogram'] < 0 and not data['golden_cross']:
        signals.append('중기하락')
    
    # 장기 신호
    if data['vs_ma200'] > 0:
        signals.append('장기상승추세')
    else:
        signals.append('장기하락추세')
    
    # 볼린저 밴드
    if data['bb_position'] < 20:
        signals.append('BB하단')
    elif data['bb_position'] > 80:
        signals.append('BB상단')
    
    return ', '.join(signals) if signals else '중립'

def get_stock_data(ticker, name, category, sector, sub_sector):
    """주가 데이터 수집 및 지표 계산"""
    try:
        stock = yf.Ticker(ticker)
        hist = stock.history(period="1y")
        
        if hist.empty or len(hist) < 2:
            return None
        
        current = hist['Close'].iloc[-1]
        prev = hist['Close'].iloc[-2] if len(hist) >= 2 else current
        
        # 수익률 계산
        change_1d = ((current / prev) - 1) * 100
        change_1w = ((current / hist['Close'].iloc[-5]) - 1) * 100 if len(hist) >= 5 else 0
        change_1m = ((current / hist['Close'].iloc[-21]) - 1) * 100 if len(hist) >= 21 else 0
        change_3m = ((current / hist['Close'].iloc[-63]) - 1) * 100 if len(hist) >= 63 else 0
        
        # 이동평균선
        ma_5 = hist['Close'].rolling(5).mean().iloc[-1] if len(hist) >= 5 else current
        ma_20 = hist['Close'].rolling(20).mean().iloc[-1] if len(hist) >= 20 else current
        ma_60 = hist['Close'].rolling(60).mean().iloc[-1] if len(hist) >= 60 else current
        ma_200 = hist['Close'].rolling(200).mean().iloc[-1] if len(hist) >= 200 else current
        
        vs_ma20 = ((current / ma_20) - 1) * 100 if ma_20 else 0
        vs_ma60 = ((current / ma_60) - 1) * 100 if ma_60 else 0
        vs_ma200 = ((current / ma_200) - 1) * 100 if ma_200 else 0
        
        # 골든크로스/데드크로스
        golden_cross = ma_20 > ma_60 if (ma_20 and ma_60) else False
        death_cross = ma_20 < ma_60 if (ma_20 and ma_60) else False
        
        # 거래량
        volume = hist['Volume'].iloc[-1]
        avg_volume = hist['Volume'].rolling(20).mean().iloc[-1] if len(hist) >= 20 else volume
        volume_ratio = (volume / avg_volume * 100) if avg_volume else 100
        
        # 기술적 지표
        rsi = calculate_rsi(hist['Close'])
        macd, macd_signal, macd_histogram = calculate_macd(hist['Close'])
        bb_position = calculate_bollinger_bands(hist['Close'])
        
        # 모멘텀 스코어 (0~100)
        momentum_score = 0
        if change_1d > 0: momentum_score += 20
        if change_1w > 0: momentum_score += 20
        if golden_cross: momentum_score += 20
        if rsi < 70 and rsi > 30: momentum_score += 20  # 과열/과매도 아님
        if macd_histogram > 0: momentum_score += 20
        
        data = {
            'name': name,
            'ticker': ticker,
            'category': category,
            'sector': sector,
            'sub_sector': sub_sector,
            'price': current,
            'change_1d': change_1d,
            'change_1w': change_1w,
            'change_1m': change_1m,
            'change_3m': change_3m,
            'vs_ma20': vs_ma20,
            'vs_ma60': vs_ma60,
            'vs_ma200': vs_ma200,
            'golden_cross': golden_cross,
            'death_cross': death_cross,
            'volume_ratio': volume_ratio,
            'rsi': rsi,
            'macd': macd,
            'macd_signal': macd_signal,
            'macd_histogram': macd_histogram,
            'bb_position': bb_position,
            'momentum_score': momentum_score,
        }
        
        data['trading_signal'] = get_trading_signal(data)
        
        return data
        
    except Exception as e:
        print(f"  ❌ {name}: {str(e)[:50]}")
        return None

print("📈 주가 데이터 수집 중...\n")

results = []
for idx, stock in enumerate(STOCKS, 1):
    print(f"[{idx}/{len(STOCKS)}] {stock['name']:20s} ... ", end='')
    data = get_stock_data(
        stock['ticker'], 
        stock['name'], 
        stock['category'],
        stock['sector'],
        stock['sub_sector']
    )
    if data:
        results.append(data)
        print("✅")
    else:
        print("❌")

print(f"\n✅ 수집 완료: {len(results)}/{len(STOCKS)}개\n")

df = pd.DataFrame(results)

# Excel 파일 생성
print("📊 Excel 파일 생성 중...\n")

now = datetime.now()
date_str = now.strftime('%Y%m%d')
time_str = now.strftime('%Y-%m-%d %H:%M')

# GitHub Actions 및 로컬 실행 모두 호환되는 경로
import os
output_dir = 'outputs'
os.makedirs(output_dir, exist_ok=True)
excel_filename = f'{output_dir}/datacenter_report_{date_str}.xlsx'

# Excel 저장
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    # 1. 종합 분석 시트
    df_export = df[[
        'name', 'ticker', 'category', 'sector', 'sub_sector',
        'price', 'change_1d', 'change_1w', 'change_1m', 'change_3m',
        'vs_ma20', 'vs_ma60', 'vs_ma200',
        'golden_cross', 'death_cross',
        'volume_ratio', 'rsi', 'macd_histogram', 'bb_position',
        'momentum_score', 'trading_signal'
    ]].copy()
    
    df_export.columns = [
        '종목명', '티커', '대분류', '중분류', '세부분류',
        '현재가', '1일수익률(%)', '1주수익률(%)', '1개월수익률(%)', '3개월수익률(%)',
        'MA20대비(%)', 'MA60대비(%)', 'MA200대비(%)',
        '골든크로스', '데드크로스',
        '거래량비율(%)', 'RSI', 'MACD히스토그램', 'BB포지션(%)',
        '모멘텀점수', '매매신호'
    ]
    
    df_export.to_excel(writer, sheet_name='종합분석', index=False)
    
    # 2. 지표 설명 시트
    indicator_info = pd.DataFrame({
        '지표명': [
            '1일수익률', '1주수익률', '1개월수익률', '3개월수익률',
            'MA20대비', 'MA60대비', 'MA200대비',
            '골든크로스', '데드크로스', '거래량비율',
            'RSI', 'MACD', 'MACD히스토그램', 'BB포지션',
            '모멘텀점수'
        ],
        '의미': [
            '전일 대비 수익률',
            '5거래일 전 대비 수익률',
            '21거래일 전 대비 수익률',
            '63거래일 전 대비 수익률',
            '20일 이동평균선 대비 현재가 위치',
            '60일 이동평균선 대비 현재가 위치',
            '200일 이동평균선 대비 현재가 위치',
            '단기 이평선이 중기 이평선을 상향돌파 (상승신호)',
            '단기 이평선이 중기 이평선을 하향돌파 (하락신호)',
            '최근 20일 평균 거래량 대비 오늘 거래량',
            '상대강도지수 (과매수/과매도 판단)',
            '이동평균수렴확산 (추세 전환 포착)',
            'MACD와 시그널선의 차이 (매매시점 판단)',
            '볼린저밴드 내 위치 (0~100, 변동성 판단)',
            '종합 모멘텀 점수 (0~100점)'
        ],
        '계산식': [
            '(현재가 / 전일종가 - 1) × 100',
            '(현재가 / 5일전종가 - 1) × 100',
            '(현재가 / 21일전종가 - 1) × 100',
            '(현재가 / 63일전종가 - 1) × 100',
            '(현재가 / MA20 - 1) × 100',
            '(현재가 / MA60 - 1) × 100',
            '(현재가 / MA200 - 1) × 100',
            'MA20 > MA60',
            'MA20 < MA60',
            '(오늘거래량 / MA20거래량) × 100',
            '100 - 100/(1+RS), RS=14일평균상승/14일평균하락',
            'EMA12 - EMA26',
            'MACD - Signal(MACD의 9일EMA)',
            '(현재가-하단밴드)/(상단밴드-하단밴드) × 100',
            '5가지 요소 합산 (각 20점)'
        ],
        '해석': [
            '+ : 상승, - : 하락',
            '+ : 상승, - : 하락',
            '+ : 상승, - : 하락',
            '+ : 상승, - : 하락',
            '+ : 이평선 위, - : 이평선 아래',
            '+ : 이평선 위, - : 이평선 아래',
            '+ : 장기상승, - : 장기하락',
            'True: 상승추세 진입',
            'True: 하락추세 진입',
            '200% 이상: 거래량 급증',
            '70이상: 과매수, 30이하: 과매도',
            '양수: 상승추세, 음수: 하락추세',
            '양수→음수: 매도신호, 음수→양수: 매수신호',
            '80이상: 상단근접, 20이하: 하단근접',
            '80점 이상: 강한 모멘텀'
        ],
        '투자활용': [
            '단기 변동성 확인',
            '단기 추세 확인',
            '중기 추세 확인',
            '장기 추세 확인',
            '단기 매매 타이밍',
            '중기 매매 타이밍',
            '장기 추세 확인',
            '매수 타이밍 포착',
            '매도 타이밍 포착',
            '관심도 상승 확인',
            '과열/침체 구간 판단',
            '추세 전환 포착',
            '구체적 매매시점',
            '단기 변동성 매매',
            '종합 매매 판단'
        ]
    })
    
    indicator_info.to_excel(writer, sheet_name='지표설명서', index=False)
    
    # 3. 대분류별 통계
    category_stats = df.groupby('category').agg({
        'change_1d': 'mean',
        'change_1w': 'mean',
        'change_1m': 'mean',
        'momentum_score': 'mean',
        'name': 'count'
    }).round(2)
    category_stats.columns = ['평균1일수익률', '평균1주수익률', '평균1개월수익률', '평균모멘텀점수', '종목수']
    category_stats.to_excel(writer, sheet_name='대분류별통계')
    
    # 4. 투자 추천 (모멘텀 상위)
    top_momentum = df.nlargest(10, 'momentum_score')[[
        'name', 'category', 'sector', 'momentum_score', 
        'change_1w', 'rsi', 'trading_signal'
    ]].copy()
    top_momentum.columns = ['종목명', '대분류', '중분류', '모멘텀점수', '1주수익률', 'RSI', '매매신호']
    top_momentum.to_excel(writer, sheet_name='투자추천TOP10', index=False)

# 서식 적용
wb = openpyxl.load_workbook(excel_filename)
ws = wb['종합분석']

# 헤더 서식
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 열 너비 자동 조정
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(excel_filename)

print(f"✅ Excel 파일 생성 완료: {excel_filename}\n")

# 텔레그램 메시지 생성 (향상된 버전)
message = "📊 데이터센터 종목 일일 리포트\n"
message += f"🕐 {time_str}\n"
message += "━━━━━━━━━━━━━━━\n\n"

# 상승 TOP 5
top_gainers = df.nlargest(5, 'change_1d')
message += "🔥 오늘 상승 TOP 5\n"
for _, row in top_gainers.iterrows():
    emoji = "🚀" if row['change_1d'] > 5 else "📈"
    message += f"{emoji} {row['name']} ({row['sub_sector']}): {row['change_1d']:+.2f}%\n"

message += "\n"

# 하락 TOP 5
top_losers = df.nsmallest(5, 'change_1d')
message += "📉 오늘 하락 TOP 5\n"
for _, row in top_losers.iterrows():
    message += f"📉 {row['name']} ({row['sub_sector']}): {row['change_1d']:+.2f}%\n"

message += "\n"

# 골든크로스
golden = df[df['golden_cross'] == True]
if len(golden) > 0:
    message += f"⭐ 골든크로스 ({len(golden)}개)\n"
    for _, row in golden.head(5).iterrows():
        message += f"• {row['name']} ({row['sub_sector']})\n"
    message += "\n"

# 모멘텀 강세 (80점 이상)
strong_momentum = df[df['momentum_score'] >= 80].nlargest(5, 'momentum_score')
if len(strong_momentum) > 0:
    message += f"💪 강한 모멘텀 ({len(df[df['momentum_score'] >= 80])}개)\n"
    for _, row in strong_momentum.iterrows():
        message += f"• {row['name']}: {row['momentum_score']:.0f}점\n"
    message += "\n"

# RSI 과매도 (매수 기회)
oversold = df[df['rsi'] < 30].sort_values('rsi')
if len(oversold) > 0:
    message += "🎯 RSI 과매도 (매수기회)\n"
    for _, row in oversold.head(3).iterrows():
        message += f"• {row['name']}: RSI {row['rsi']:.1f}\n"
    message += "\n"

# 거래량 급증
volume_spike = df[df['volume_ratio'] > 200].nlargest(5, 'volume_ratio')
if len(volume_spike) > 0:
    message += "📊 거래량 급증 (평균 대비 2배↑)\n"
    for _, row in volume_spike.iterrows():
        message += f"• {row['name']}: {row['volume_ratio']:.0f}%\n"
    message += "\n"

# 대분류별 현황
message += "━━━━━━━━━━━━━━━\n"
message += "📂 대분류별 현황\n"
for category in df['category'].unique():
    cat_df = df[df['category'] == category]
    up = len(cat_df[cat_df['change_1d'] > 0])
    total = len(cat_df)
    avg_change = cat_df['change_1d'].mean()
    message += f"• {category}: {up}/{total}개 상승 (평균 {avg_change:+.2f}%)\n"

message += "\n"

# 전체 통계
up_count = len(df[df['change_1d'] > 0])
down_count = len(df[df['change_1d'] < 0])

message += "━━━━━━━━━━━━━━━\n"
message += f"📈 상승: {up_count}개\n"
message += f"📉 하락: {down_count}개\n"
message += f"📊 총 {len(results)}개 종목\n"
message += f"📁 Excel: datacenter_report_{date_str}.xlsx"

print("📱 텔레그램 전송 중...\n")

# 1. 텍스트 메시지 전송
url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message}

try:
    response = requests.post(url, data=payload)
    if response.status_code == 200:
        print("✅ 텔레그램 메시지 전송 성공!")
    else:
        print(f"❌ 메시지 전송 실패: {response.status_code}")
except Exception as e:
    print(f"❌ 메시지 전송 오류: {e}")

# 2. Excel 파일 전송
print("📎 Excel 파일 전송 중...\n")
file_url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

try:
    with open(excel_filename, 'rb') as file:
        files = {'document': file}
        data = {
            'chat_id': TELEGRAM_CHAT_ID,
            'caption': f'📊 데이터센터 일일 리포트\n📅 {now.strftime("%Y-%m-%d %H:%M")}'
        }
        response = requests.post(file_url, files=files, data=data)
        
        if response.status_code == 200:
            print("✅ Excel 파일 전송 성공!")
        else:
            print(f"❌ 파일 전송 실패: {response.status_code}")
            print(f"   (Artifacts에서 다운로드 가능)")
except Exception as e:
    print(f"❌ 파일 전송 오류: {e}")
    print(f"   (Artifacts에서 다운로드 가능)")

print("\n" + "="*70)
print("✅ 작업 완료!")
print(f"📊 Excel 파일: {excel_filename}")
print("="*70)
